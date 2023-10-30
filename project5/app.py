import io
import os
import openpyxl
import msoffcrypto
import pandas as pd

from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from flask import Flask, request, render_template, redirect, url_for



app=Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['AUDIT_FOLDER'] = 'audit'




@app.route('/')
def index():
    return render_template('index.html')  
    
   
decrypted_workbook=io.BytesIO()
@app.route('/upload', methods=['GET', 'POST'])
def upload_file():

    lst_sheet=[]
    worksheet_names=None  # Initialize worksheet_names
    
    
    if request.method == 'POST':
        if 'file' not in request.files:
            return "No file part"
        
        file = request.files['file']
        if file.filename == '':
            return "No selected file"
        
        if file:
            filename=secure_filename(file.filename)
            filepath=os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            #try:
            #    workbook=pd.ExcelFile(filepath)
            #    worksheet_names=workbook.sheet_names
                
            try:
                with open(filepath, 'rb') as file:
                    office_file=msoffcrypto.OfficeFile(file)
                    office_file.load_key(password='')
                    office_file.decrypt(decrypted_workbook)
                    
                worksheet_names=openpyxl.load_workbook(filename=decrypted_workbook)
                for sheet in worksheet_names.sheetnames:
                    lst_sheet.append(sheet)
                    
                
            except Exception as e:
                return f"An error occurred while reading the file: {str(e)}"
                       
            return render_template('upload_single.html', worksheet_names=lst_sheet, uploaded_file=decrypted_workbook)

    return render_template('upload_single.html', worksheet_names=lst_sheet, uploaded_file=None)


@app.route('/fetch_data', methods=['POST'])
def fetch_data():
    worksheet_name=request.form['worksheet']
    uploaded_file=request.form['uploaded_file']

    try:
        #filepath=os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file)   
        #df=pd.read_excel(filepath, sheet_name=worksheet_name)
        
        df=pd.read_excel(decrypted_workbook, sheet_name=worksheet_name, engine='openpyxl')                

        # Debugging statements
        print(f"worksheet_name: {worksheet_name}")
        print(f"uploaded_file: {uploaded_file}")
        print(df.head(10))

        # Convert the DataFrame to an HTML table
        html_table=df.to_html(classes="table table-striped")
        

        return html_table

    except Exception as e:
        error_message = f"An error occurred while reading the worksheet: {str(e)}"
        return error_message, 500  # Return a 500 Internal Server Error status code
        

@app.route('/get_parameters', methods=['GET', 'POST'])
def get_parameters():

    if request.method == 'POST':
        if 'file' not in request.files:
            return "No file part"
        
        file = request.files['file']
        if file.filename == '':
            return "No selected file"
    select_quarter=request.form['quarter']
    select_month=request.form['month']







if __name__ == '__main__':
    app.run(debug=True)
