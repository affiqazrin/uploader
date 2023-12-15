import io
import os

import glob
import openpyxl
import pandas as pd
import sqlalchemy as sa

import ipywidgets as widgets
from ipywidgets import interactive

from urllib.parse import quote  
from sqlalchemy.engine import create_engine

import datetime
from datetime import date
from datetime import datetime
from random import randint, randrange

import msoffcrypto

class StatementProcessor:
    def __init__(self, db_connection_string):
        self.engine=db_connection_string
        self.df={}
        self.params_cols={
            'insx_month': 'insx_month', 
            'insx_cycle': 'insx_cycle', 
            'insx_version': 'insx_version', 
            'insg_src_file': 'insx_src_file'
        }
        
        #print(self.engine)
    
    def modification_date(self, insx_file_name):
        t=os.path.getmtime(insx_file_name)
        return datetime.fromtimestamp(t).strftime('%y%m%d')
        
    def read_insx_file(self, f, decrypted_workbook):        
        df={}
        lst_sheet=[]
        #decrypted_workbook=io.BytesIO()
        #for f in glob.glob(os.path.join(INFILE, insx_file_name)):
        #   print(os.path.basename(f))
        
        try:
            #print('pwd')
            with open(f, 'rb') as file:
                office_file=msoffcrypto.OfficeFile(file)
                office_file.load_key(password='prm83881')
                office_file.decrypt(decrypted_workbook)

            workbook=openpyxl.load_workbook(filename=decrypted_workbook)
            #print(workbook.sheetnames)
            try:
                for sheet in workbook.sheetnames:
                    #print(sheet)
                    lst_sheet.append(sheet)
                    df[f'{sheet}']=pd.read_excel(decrypted_workbook, sheet_name=sheet, engine='openpyxl')
                    #df[f'{sheet}']=df[f'{sheet}'].loc[:, ~df[f'{sheet}'].columns.str.contains('^Unnamed', na=False)] 
            except Exception as b:
                print(b)
                pass

        except Exception as a:
            #print('no pwd ', a)
            xl=pd.ExcelFile(f)
            #print(xl.sheet_names)
            try:
                for sheet in xl.sheet_names:
                    #print(sheet)
                    lst_sheet.append(sheet)
                    df[f'{sheet}']=pd.read_excel(xl, sheet_name=sheet,)
                    #df[f'{sheet}']=df[f'{sheet}'].loc[:, ~df[f'{sheet}'].columns.str.contains('^Unnamed', na=False)]
            except Exception as c:
                print(c)
                pass
        
        print(f'insx_version: {self.modification_date(f)}_{randint(100, 999)}')
        #file.close()       

        return df, lst_sheet, f
    
    def read_insx_config(self, f, insg_cycle):
        insx_file_name=os.path.basename(f).split('.')[0]
        sql=f"""
            SELECT * FROM insg_statement_config WHERE insg_calculation_file_name='{insx_file_name.replace('_', ' ')}' AND insg_cycle='{insg_cycle}'
            """
        print(f"Executing SQL query: {sql}")
        
        insg_to_insx=pd.read_sql(sql, self.engine).iloc[0].to_dict()
        insg_to_insx['insx_version']=f'{self.modification_date(f)}_{randint(100, 999)}'
        
        del insg_to_insx['insg_last_update']
               
        return insg_to_insx
    
    def process_data(self, df, insg_to_insx, sheet, cycle, month):
        header_row=int(insg_to_insx['insg_xls_start_row'])
        #print(insg_to_insx)

        # Transform data
        df.columns=df.iloc[header_row]
        df_cleaned=df.iloc[header_row+1:].reset_index(drop=True)
        print(df_cleaned)
        
        # Process insx table
        df_cleaned=df_cleaned.loc[:, df_cleaned.columns.notna()]
        df_cleaned=df_cleaned.dropna(axis=1, how='all')  # Drop columns with all NaN values
        df_filtered=df_cleaned[df_cleaned.insx_staff_id.str.contains(r'^100[0-9]{5}$', regex=True, na=False)]             
        df_filtered.rename(columns=self.params_cols, inplace=True)
        
        # Rename columns
        for col in self.params_cols:
            #print(col, params_cols[col])
            # Add new columns if not exist
            if self.params_cols[col] not in df_filtered.columns:
                df_filtered[self.params_cols[col]]=None
                
        # Convert value column to float data type
        convert_lst=['RP', 'rp' 'RM', 'rm', 'CC', 'cc', 'tgt', 'ach', 'act', 'rate', 'payout', 
                     'ye_rate', 'amt', 'amount', 'deduct', 'rem1','rem2', 'rem3', 'rem4', 'total',
                     'rm', 'ret', 'oth_clw', 'cf_clw']
        
        for i in range(len(convert_lst)):
            #print(convert_lst[i])
            for cols in df_filtered.columns:
                if convert_lst[i] in cols:
                    df_filtered[cols]=df_filtered[cols].astype(float)
           
        self.params_cols={
            'insx_cycle': cycle,
            'insx_month': month,
            'insx_version': insg_to_insx['insx_version'],
            'insx_row_count': df_filtered.shape[0],
            'insx_timestamp_upload': ''
        }

        # Update value in new added columns
        df_filtered['insx_month']=self.params_cols['insx_month']
        df_filtered['insx_cycle']=self.params_cols['insx_cycle']
        df_filtered['insx_version']=self.params_cols['insx_version']
        df_filtered['insx_src_file']=insg_to_insx['insg_src_file']

        #print(df_filtered.head())        
        self.df=df_filtered
                   
        return df_filtered
    
    def process_statement(self, f, insg_cycle, cycle, month):        
        insg_to_insx=self.read_insx_config(f, insg_cycle)        
        print(self.df.head(10))

        try:           
            # Process insx table
            insg_table_name=insg_to_insx['insg_table_src']
            print(f"Inserting data into table: {insg_table_name}")            
            self.df.to_sql(con=self.engine, name=insg_table_name, if_exists='append', index=False)
            print("Data inserted successfully!")
            
        except Exception as e:
            print(f"Error processing statement: {e}")
            
        try:
            # Call stored procedure
            sp_query=f"""
                CALL {insg_to_insx['insg_sp']}('{self.params_cols['insx_month']}', '{insg_to_insx['insg_cycle']}', '{self.params_cols['insx_version']}', '{insg_to_insx['insg_src_file']}')
                """
            print(f"Executing stored procedure:\n{sp_query}")
            with self.engine.begin() as connection:
                connection.execute(sp_query)
            print("Stored procedure executed successfully!")

            # Get timestamp statement upload
            timestamp_query=f"""
                SELECT DISTINCT(inst_insx_version) FROM inst_statement WHERE inst_month='{self.params_cols['insx_month']}' AND inst_src_file='{insg_to_insx['insg_src_file']}' AND '{self.params_cols['insx_version']}'
                """
            print(f"Executing timestamp query: {timestamp_query}")
            timestamp_check=pd.read_sql(timestamp_query, self.engine)
            try:
                self.params_cols['insx_timestamp_upload']=str(datetime.now().strftime("%m/%d/%Y, %H:%M:%S")) if timestamp_check['inst_insx_version'][0] == self.params['insx_version'] else ''
            except Exception as e:
                print(e)
            print("Statement processed successfully!")
            print(self.params_cols.iloc[0].to_dict())
            
        except Exception as e:
            print(f"Error processing statement: {e}")


#processed_df=processor.process_statement(self, f, sheet, insg_cycle, cycle, month)
#processed_df

