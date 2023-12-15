import io
import os
import openpyxl
import msoffcrypto
import pandas as pd

from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from sqlalchemy import create_engine, text
from flask import Flask, request, render_template, redirect, url_for


app=Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config.from_object('config.TestConfig')  # Change to 'config.ProductionConfig' for production

# Initialize the SQLAlchemy extension
db = SQLAlchemy(app)


def create_db_engine():
    try:
        # Create an SQLAlchemy engine for use with Pandas
        #engine = create_engine('sqlite:///titanic.db')
        engine = create_engine(app.config['SQLALCHEMY_DATABASE_URI'])
        print(engine)
        return engine

    except Exception as e:
        print(f"Error creating database engine: {str(e)}")
        return None



@app.route('/')
def index():
    return render_template('index.html')

decrypted_workbook=io.BytesIO()
@app.route('/upload', methods=['GET', 'POST'])
def upload_file():

    lst_sheet=[]
    worksheet_names=None  # Initialize worksheet_names
    
    
    if request.method == 'POST':
        if 'file' not in request.files:
            return "No file part"
        
        file = request.files['file']
        if file.filename == '':
            return "No selected file"
        
        if file:
            filename=secure_filename(file.filename)
            filepath=os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            #try:
            #    workbook=pd.ExcelFile(filepath)
            #    worksheet_names=workbook.sheet_names
                
            try:
                with open(filepath, 'rb') as file:
                    office_file=msoffcrypto.OfficeFile(file)
                    office_file.load_key(password='prm83881')
                    office_file.decrypt(decrypted_workbook)
                    
                worksheet_names=openpyxl.load_workbook(filename=decrypted_workbook)
                for sheet in worksheet_names.sheetnames:
                    lst_sheet.append(sheet)
                                    
            except Exception as e:
                return f"An error occurred while reading the file: {str(e)}"
                       
            return render_template('upload_single.html', worksheet_names=lst_sheet, uploaded_file=decrypted_workbook)

    return render_template('upload_single.html', worksheet_names=lst_sheet, uploaded_file=None)

@app.route('/fetch_data', methods=['POST'])
def fetch_data():
    worksheet_name=request.form['worksheet']
    uploaded_file=request.form['uploaded_file']

    try:
        #filepath=os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file)   
        #df=pd.read_excel(filepath, sheet_name=worksheet_name)
        
        df=pd.read_excel(decrypted_workbook, sheet_name=worksheet_name, engine='openpyxl')                

        # Debugging statements
        print(f"worksheet_name: {worksheet_name}")
        print(f"uploaded_file: {uploaded_file}")
        print(df.head(10))

        # Convert the DataFrame to an HTML table
        html_table=df.to_html(classes="table table-striped")
        

        return html_table

    except Exception as e:
        error_message = f"An error occurred while reading the worksheet: {str(e)}"
        return error_message, 500  # Return a 500 Internal Server Error status code

sql_query=None             
@app.route('/get_data', methods=['GET', 'POST'])
def get_data():
    engine = create_db_engine()
    print(engine)
    if engine is None:
        return "Error creating database engine", 500

    # Sample SQL query, replace with your own query    
    # Get the SQL query from the form
    # sql_query = "SELECT * FROM 'Observation' LIMIT 10"
    
    if request.method == 'POST':
        sql_query = request.form['query']
        print(sql_query)
       

        try:
            # Fetch data from the MySQL database and convert it to a Pandas DataFrame
            with db.engine.begin() as conn:
                result = conn.execute(text(sql_query))
            df=pd.DataFrame(result)
        
            #df = pd.read_sql(sql_query, engine.raw_connection())
            #df=pd.DataFrame(engine.connect().execute(text(sql_query)))
            print(df.head(10))

            # Convert the DataFrame to HTML
            table_html = df.head(10).to_html(classes='table table-bordered', index=False)

            # Render the 'results.html' template with the HTML representation of the DataFrame
            return render_template('results.html', table=table_html)
            
        except Exception as e:
            print(f"Error fetching data from the database: {str(e)}")
            return "Error fetching data from the database", 500
            
    return render_template('results.html')
       
@app.route('/list', methods=['GET', 'POST'])
def etl_job():
    return render_template('list.html')




if __name__ == '__main__':
    app.run(debug=True)